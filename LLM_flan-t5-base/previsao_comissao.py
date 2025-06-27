# previsao_comissao.py

# Previsão de Comissão com Modelos Estatísticos e LLM Local

"""
Este script realiza análises preditivas de valor de comissão a partir de uma base fictícia de apólices,
utilizando modelos estatísticos (Regressão Linear e Random Forest) e uma LLM local (FLAN-T5) para explicações.

Etapas:
1. Carregamento e pré-processamento dos dados
2. Modelagem estatística com Regressão Linear e Random Forest (com divisão treino/teste)
3. Geração de previsões para os próximos 30 dias
4. Geração de gráficos de histórico (últimos 12 meses) + previsão
5. Geração de aba explicativa com métricas estatísticas de performance do modelo (com LLM)
6. Salvamento dos resultados em arquivo Excel

Requisitos:
- pip install pandas scikit-learn transformers torch openpyxl matplotlib
"""

# === CONFIGURAÇÃO INICIAL =====================================================
# Define o diretório de entrada onde os arquivos de dados (Excel) estão localizados.
INPUT_DIR = r'INSERIR_CAMINHO_AQUI' # Substitua pelo caminho real dos arquivos de entrada.
# Define o diretório de saída onde os resultados (Excel e gráficos) serão salvos.
OUTPUT_DIR = r'INSERIR_CAMINHO_AQUI' # Substitua pelo caminho real do diretório de saída.
# Nome do modelo de Linguagem Grande (LLM) a ser carregado para gerar explicações.
MODEL_NAME = "google/flan-t5-base"

# =============================================================================
# Importa módulos e bibliotecas necessários para o script.
import os # Para interagir com o sistema operacional (caminhos, diretórios).
import sys # Para interagir com o interpretador Python.
import time # Para medir o tempo de execução.
import getpass # Para obter o nome de usuário do sistema.
import platform # Para obter informações sobre o sistema operacional.
import warnings # Para gerenciar avisos.
from glob import glob # Para encontrar arquivos que correspondem a um padrão específico.
from datetime import datetime, timedelta # Para manipulação de datas e tempo.
import argparse # Para lidar com argumentos de linha de comando.

import pandas as pd # Para manipulação e análise de dados tabulares (DataFrames).
import numpy as np # Para operações numéricas e arrays.
from sklearn.model_selection import train_test_split # Para dividir dados em conjuntos de treino e teste.
from sklearn.linear_model import LinearRegression # Modelo de Regressão Linear.
from sklearn.ensemble import RandomForestRegressor # Modelo de Random Forest para regressão.
# IMPORTANTE: A linha abaixo foi ajustada para incluir 'root_mean_squared_error'
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score, root_mean_squared_error # Métricas de avaliação de modelos.
from transformers import AutoTokenizer, AutoModelForSeq2SeqLM # Para carregar o tokenizer e o modelo LLM.
import torch # Biblioteca PyTorch, usada pelo Hugging Face Transformers para LLM.
import matplotlib.pyplot as plt # Para criação de gráficos.
# Importações adicionais para formatação Excel
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configurações para o Matplotlib para padronizar a aparência dos gráficos.
plt.style.use('seaborn-v0_8-darkgrid') # Define um estilo visual para os gráficos.
plt.rcParams['figure.figsize'] = (12, 6) # Define o tamanho padrão das figuras (largura, altura).
plt.rcParams['lines.linewidth'] = 2 # Define a largura padrão das linhas nos gráficos.
plt.rcParams['font.size'] = 12 # Define o tamanho da fonte padrão para o texto nos gráficos.

# Ignora avisos específicos da biblioteca scikit-learn para evitar poluição no console.
warnings.filterwarnings("ignore", category=UserWarning, module="sklearn")

# Configura o parser de argumentos de linha de comando.
parser = argparse.ArgumentParser(add_help=False) # 'add_help=False' evita que o parser adicione a opção -h/--help automaticamente.
# Adiciona o argumento '-i' ou '--input' para especificar o diretório de entrada.
parser.add_argument("-i", "--input", dest="cli_input")
# Adiciona o argumento '-o' ou '--output' para especificar o diretório de saída.
parser.add_argument("-o", "--output", dest="cli_output")
args, _ = parser.parse_known_args() # Analisa os argumentos fornecidos na linha de comando.

# Sobrescreve INPUT_DIR e OUTPUT_DIR se forem fornecidos via linha de comando.
if args.cli_input:
    INPUT_DIR = args.cli_input
if args.cli_output:
    OUTPUT_DIR = args.cli_output

# Verifica se o diretório de entrada existe; caso contrário, levanta um erro.
if not os.path.isdir(INPUT_DIR):
    raise FileNotFoundError(f"Pasta de entrada não encontrada: {INPUT_DIR}")
# Cria o diretório de saída se ele não existir (exist_ok=True evita erro se já existir).
os.makedirs(OUTPUT_DIR, exist_ok=True)
# Define o caminho para o subdiretório onde os gráficos serão salvos.
GRAFICOS_DIR = os.path.join(OUTPUT_DIR, "graficos")
# Cria o diretório para gráficos se ele não existir.
os.makedirs(GRAFICOS_DIR, exist_ok=True)

# Imprime informações sobre o ambiente de execução e os diretórios.
print("\n🧑 Usuário:", getpass.getuser()) # Mostra o nome do usuário logado.
print("💻 Máquina:", platform.node()) # Mostra o nome da máquina.
print("🐍 Python :", sys.version.split()[0]) # Mostra a versão do Python.
print("🗕️ Início  :", datetime.now().strftime("%Y-%m-%d %H:%M:%S")) # Mostra a data e hora de início.
print("\nInput :", INPUT_DIR) # Confirma o diretório de entrada.
print("Output:", OUTPUT_DIR) # Confirma o diretório de saída.
print("="*60, "\n") # Imprime uma linha de separação.

# === 1. Carregar e Pré-processar Dados =======================================
print("🔍 Varredura de .xlsx…")
# Procura por todos os arquivos .xlsx no diretório de entrada.
arquivos_xlsx = glob(os.path.join(INPUT_DIR, "*.xlsx"))
# Se nenhum arquivo .xlsx for encontrado, levanta um erro.
if not arquivos_xlsx:
    raise FileNotFoundError("Nenhum .xlsx encontrado na pasta de entrada.")
# Imprime o nome de cada arquivo .xlsx encontrado.
for arq in arquivos_xlsx:
    print("   •", os.path.basename(arq))

# Lê cada arquivo Excel e armazena seus DataFrames em uma lista.
lista_df = [pd.read_excel(arq) for arq in arquivos_xlsx]
# Concatena todos os DataFrames em um único DataFrame.
df_original = pd.concat(lista_df, ignore_index=True)

# Define as colunas que são obrigatórias no DataFrame.
colunas_obrigatorias = ["Data de Emissão", "Valor Comissão", "Seguradora", "Produto"]
# Verifica se todas as colunas obrigatórias estão presentes. Se não, levanta um erro.
if not all(col in df_original.columns for col in colunas_obrigatorias):
    raise ValueError(f"As colunas {colunas_obrigatorias} são obrigatórias.")

# Converte a coluna 'Data de Emissão' para o tipo datetime.
df_original["Data de Emissão"] = pd.to_datetime(df_original["Data de Emissão"])
# Converte as colunas 'Seguradora' e 'Produto' para o tipo string.
df_original["Seguradora"] = df_original["Seguradora"].astype(str)
df_original["Produto"] = df_original["Produto"].astype(str)
# Converte 'Valor Comissão' para numérico, tratando erros (valores inválidos viram NaN).
df_original["Valor Comissão"] = pd.to_numeric(df_original["Valor Comissão"], errors='coerce')
# Remover as linhas onde 'Valor Comissão' é NaN (nulo) após a conversão.
df_original.dropna(subset=["Valor Comissão"], inplace=True)

print(f"✅ Dados carregados. Total de {len(df_original)} registros.")

# === Carregar LLM para Geração de Explicações ================================
print(f"🧠 Carregando LLM ({MODEL_NAME}) para geração de explicações...")
# Carrega o tokenizer (responsável por converter texto em tokens que o modelo entende) do modelo especificado.
tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME)
# Carrega o modelo de Linguagem Grande (LLM) pré-treinado.
model_llm = AutoModelForSeq2SeqLM.from_pretrained(MODEL_NAME)
# Verifica se uma GPU (CUDA) está disponível e move o modelo para a GPU, se sim, para acelerar o processamento.
if torch.cuda.is_available():
    model_llm.to('cuda')
    print("   • LLM carregada para GPU.")
else:
    print("   • LLM carregada para CPU.")

# Define uma função para gerar explicações usando a LLM.
def generate_llm_explanation(prompt_text):
    # Tokeniza o texto de entrada (prompt), convertendo-o em tensores para o modelo.
    # 'max_length' e 'truncation' garantem que o texto não exceda o limite do modelo.
    inputs = tokenizer(prompt_text, return_tensors="pt", max_length=512, truncation=True)
    # Se a GPU estiver disponível, move os inputs para a GPU.
    if torch.cuda.is_available():
        inputs = {k: v.to('cuda') for k, v in inputs.items()}
    # Desativa o cálculo de gradientes para economizar memória e acelerar a inferência.
    with torch.no_grad():
        # Gera a saída do modelo com base nos inputs.
        outputs = model_llm.generate(
            **inputs,
            max_new_tokens=100, # Limita o número de novos tokens gerados na resposta.
            num_beams=5, # Usa busca em feixe para gerar respostas mais coerentes.
            early_stopping=True # Para a geração se o modelo determinar que já encontrou uma boa resposta.
        )
    # Decodifica os tokens gerados de volta para texto legível, ignorando tokens especiais.
    return tokenizer.decode(outputs[0], skip_special_tokens=True)

# === 2. Modelagem e Previsões para os Próximos 30 Dias =======================
print("\n🤖 Gerando previsões estatísticas para os próximos 30 dias…")
resultados = [] # Lista para armazenar os resultados das previsões.
metricas_modelos = [] # Lista para armazenar as métricas de desempenho dos modelos.
# CORREÇÃO: data_inicio_exec movida para fora do loop para capturar o tempo total do processo
data_inicio_exec = time.time() # Registra o tempo de início para calcular a duração.

# Itera sobre cada grupo único de "Seguradora" e "Produto" no DataFrame original.
for (seg, prod), grupo in df_original.groupby(["Seguradora", "Produto"]):
    # Pula grupos com menos de 2 registros, pois não é possível treinar um modelo com tão poucos dados.
    if len(grupo) < 2:
        print(f"   • Pulando {seg} - {prod}: Poucos dados para modelagem ({len(grupo)} registros).")
        continue

    # Cria uma cópia do grupo e o ordena pela 'Data de Emissão'.
    grupo = grupo.sort_values("Data de Emissão").copy()
    # Adiciona uma coluna 'Dias' representando o número de dias desde a primeira emissão do grupo.
    grupo["Dias"] = (grupo["Data de Emissão"] - grupo["Data de Emissão"].min()).dt.days

    # Define as features (variáveis independentes) e o target (variável dependente).
    X_full = grupo[["Dias"]] # 'Dias' é a feature.
    y_full = grupo["Valor Comissão"] # 'Valor Comissão' é o target.

    # Divisão treino/teste para avaliação do modelo
    # Divide os dados, usando 80% para treino e 20% para teste, mantendo a ordem temporal.
    split_point = int(len(grupo) * 0.8)
    X_train, X_test = X_full.iloc[:split_point], X_full.iloc[split_point:]
    y_train, y_test = y_full.iloc[:split_point], y_full.iloc[split_point:]

    # Pula se a divisão resultar em conjuntos de treino ou teste vazios.
    if len(X_train) == 0 or len(X_test) == 0:
        print(f"   • Pulando {seg} - {prod}: Divisão treino/teste resultou em conjuntos vazios.")
        continue

    # Instancia os modelos de regressão.
    lr = LinearRegression() # Regressão Linear.
    rf = RandomForestRegressor(random_state=42) # Random Forest Regressor com uma semente para reprodutibilidade.

    try:
        # Treina os modelos com os dados de treino.
        lr.fit(X_train, y_train)
        rf.fit(X_train, y_train)

        # Avalia os modelos no conjunto de teste.
        y_pred_lr_test = lr.predict(X_test) # Previsões do modelo de Regressão Linear.
        y_pred_rf_test = rf.predict(X_test) # Previsões do modelo de Random Forest.
        # Calcula a média das previsões de ambos os modelos para uma previsão combinada.
        y_pred_avg_test = (y_pred_lr_test + y_pred_rf_test) / 2

        # Calcula as métricas de erro para a previsão combinada.
        mae_test = mean_absolute_error(y_test, y_pred_avg_test) # Erro Absoluto Médio.
        # AJUSTE: Usando root_mean_squared_error diretamente (para scikit-learn >= 1.2)
        rmse_test = root_mean_squared_error(y_test, y_pred_avg_test) # Raiz do Erro Quadrático Médio.
        r2_test = r2_score(y_test, y_pred_avg_test) # Coeficiente de Determinação (R²).

        # Armazena as métricas de desempenho do modelo.
        metricas_modelos.append({
            "Seguradora": seg,
            "Produto": prod,
            "MAE (Modelo)": round(mae_test, 2),
            "RMSE (Modelo)": round(rmse_test, 2),
            "R² (Modelo)": round(r2_test, 4)
        })

        # Previsões para os próximos 30 dias (usando o modelo treinado com TODOS os dados históricos)
        # É importante retreinar com todos os dados para a previsão futura mais precisa
        lr_final = LinearRegression().fit(X_full, y_full) # Retreina LR com todos os dados.
        rf_final = RandomForestRegressor(random_state=42).fit(X_full, y_full) # Retreina RF com todos os dados.

        ultima_data_historica = grupo["Data de Emissão"].max() # Última data de emissão no histórico.
        data_min_historica = grupo["Data de Emissão"].min() # Primeira data de emissão no histórico.

        # Loop para prever os próximos 30 dias.
        for i in range(1, 31):
            data_prevista = ultima_data_historica + timedelta(days=i) # Calcula a data futura.
            dias_futuros = (data_prevista - data_min_historica).days # Calcula os dias desde a data mínima histórica.
            dias_df = pd.DataFrame({"Dias": [dias_futuros]}) # Cria um DataFrame com os dias futuros para previsão.

            prev_lr = lr_final.predict(dias_df)[0] # Previsão da Regressão Linear.
            prev_rf = rf_final.predict(dias_df)[0] # Previsão do Random Forest.
            # Calcula a média das previsões, garantindo que o valor da comissão não seja negativo.
            valor_estimado = max(0, (prev_lr + prev_rf) / 2)

            # Adiciona os resultados da previsão à lista.
            resultados.append({
                "Seguradora": seg,
                "Produto": prod,
                "Data": data_prevista.date(), # Armazena apenas a data.
                "Previsão Comissão": valor_estimado # Armazena como float para cálculos futuros.
            })
    # Captura e imprime quaisquer erros que ocorram durante a modelagem.
    except Exception as e:
        print(f"   • Erro ao modelar {seg} - {prod}: {e}")
        # Adiciona métricas nulas caso ocorra um erro na modelagem.
        metricas_modelos.append({
            "Seguradora": seg, "Produto": prod,
            "MAE (Modelo)": None, "RMSE (Modelo)": None, "R² (Modelo)": None
        })

# Converte as listas de resultados e métricas em DataFrames.
df_resultados = pd.DataFrame(resultados)
df_metricas_modelos = pd.DataFrame(metricas_modelos)

# === 3. Salvar Excel de Resultados ===========================================
print("\n📄 Salvando Excel de previsões e métricas…")

excel_saida = os.path.join(OUTPUT_DIR, "previsao_llm_30dias.xlsx") # Define o caminho do arquivo Excel de saída.

# Não formatamos df_resultados para string aqui, ele será salvo como número/data e formatado via openpyxl
if df_resultados.empty:
    print("⚠️ Atenção: DataFrame de resultados vazio. A aba 'Previsao_30_Dias' será gerada vazia.")
    df_resultados_formatado = pd.DataFrame(columns=["Seguradora", "Produto", "Data", "Previsão Comissão"]) # Cria um DF vazio com as colunas esperadas
else:
    # Usaremos df_resultados diretamente, sem df_resultados_formatado para manter tipos numéricos
    df_resultados_formatado = df_resultados.copy() # Criamos uma cópia só para consistência com o nome da variável se você preferir

# Usa pd.ExcelWriter para criar e salvar o arquivo Excel.
with pd.ExcelWriter(excel_saida, engine="openpyxl") as writer:
    # Salva o DataFrame de resultados na aba "Previsao_30_Dias".
    df_resultados_formatado.to_excel(writer, sheet_name="Previsao_30_Dias", index=False)
    # A aba de métricas será adicionada posteriormente com as explicações da LLM.

# === 4. Geração de Gráficos ==================================================
print("\n📊 Gerando gráficos por Seguradora e Produto…")

# Itera sobre cada grupo de "Seguradora" e "Produto" no DataFrame de resultados de previsão.
# Adicionada uma verificação de 'df_resultados.empty' para evitar erro se não houver previsões
if not df_resultados.empty:
    for (seg, prod), grupo_prev in df_resultados.groupby(["Seguradora", "Produto"]):
        historico = df_original[(df_original["Seguradora"] == seg) & (df_original["Produto"] == prod)].copy()
        historico_grouped = historico.groupby("Data de Emissão")["Valor Comissão"].sum().reset_index()

        # Filtrar histórico para os últimos 12 meses
        # Considera 12m a partir da data atual de execução, para o histórico
        corte_data = datetime.now() - pd.DateOffset(months=12)
        historico_grouped = historico_grouped[historico_grouped["Data de Emissão"] >= corte_data]

        plt.figure(figsize=(12, 6))
        plt.plot(historico_grouped["Data de Emissão"], historico_grouped["Valor Comissão"],
                 label="Histórico (últ. 12m)", marker="o", markersize=4, color='blue')
        plt.plot(grupo_prev["Data"], grupo_prev["Previsão Comissão"],
                 label="Previsão (30d)", linestyle="--", marker="x", markersize=4, color='red')

        plt.title(f"Previsão de Comissão: {seg} – {prod}", fontsize=16)
        plt.xlabel("Data", fontsize=14)
        plt.ylabel("Comissão (R$)", fontsize=14)
        plt.legend(fontsize=12)
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.xticks(rotation=45)
        plt.tight_layout()

        nome_fig = f"{seg}_{prod}.png".replace(" ", "_").replace("/", "_").replace("\\", "_")
        plt.savefig(os.path.join(GRAFICOS_DIR, nome_fig))
        plt.close()
    print(f"✅ Gráficos gerados e salvos em: {GRAFICOS_DIR}")
else:
    print("🚫 Não há dados de previsão para gerar gráficos. O DataFrame de resultados está vazio.")

# === 5. Geração de Explicações Estatísticas com LLM ==========================
print("\n🧠 Gerando aba de explicações estatísticas com LLM...")

explicacoes = [] # Lista para armazenar as explicações geradas pela LLM.
# Itera sobre cada linha do DataFrame de métricas do modelo.
# Adicionada uma verificação de 'df_metricas_modelos.empty' para evitar erro se não houver métricas
if not df_metricas_modelos.empty:
    for index, row in df_metricas_modelos.iterrows():
        seg = row["Seguradora"]
        prod = row["Produto"]
        mae = row["MAE (Modelo)"]
        rmse = row["RMSE (Modelo)"]
        r2 = row["R² (Modelo)"]

        interpretação_llm = "Não foi possível gerar uma interpretação." # Mensagem padrão caso não haja interpretação.

        # Gera a interpretação apenas se o R² não for nulo (ou seja, se o modelo foi treinado com sucesso).
        if pd.notna(r2):
            # Constrói o prompt para a LLM, incluindo as métricas do modelo.
            prompt = (
                f"Considere as seguintes métricas de um modelo de previsão de comissão "
                f"para a seguradora '{seg}' e produto '{prod}':\n"
                f"- Erro Absoluto Médio (MAE): R${mae:,.2f}\n"
                f"- Raiz do Erro Quadrático Médio (RMSE): R${rmse:,.2f}\n"
                f"- Coeficiente de Determinação (R²): {r2:.4f}\n\n"
                f"Forneça uma breve interpretação dessas métricas, explicando o que elas significam "
                f"em termos da capacidade do modelo de prever o valor de comissão."
            )
            # Chama a função para gerar a explicação usando a LLM.
            interpretação_llm = generate_llm_explanation(prompt)

        # Adiciona as métricas e a interpretação da LLM à lista de explicações.
        explicacoes.append({
            "Seguradora": seg,
            "Produto": prod,
            "MAE (Modelo)": mae,
            "RMSE (Modelo)": rmse,
            "R² (Modelo)": r2,
            "Interpretação da LLM": interpretação_llm
        })

    # Converte a lista de explicações em um DataFrame.
    df_exp = pd.DataFrame(explicacoes)

    # Salva o DataFrame de explicações em uma nova aba no arquivo Excel existente.
    print("📄 Salvando aba 'Resumo_Explicativo' no Excel...")
    # Abre o arquivo Excel em modo de anexar ('a') e substitui a aba se já existir.
    with pd.ExcelWriter(excel_saida, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_exp.to_excel(writer, sheet_name="Resumo_Explicativo", index=False)
else:
    print("🚫 Não há métricas de modelo para gerar explicações. A aba 'Resumo_Explicativo' não será criada ou será vazia.")

# === 6. Aplicação de Formatação Final no Excel (PÓS-SALVAMENTO) ==============
print("\n📝 Aplicando formatação final ao arquivo Excel...")
try:
    wb = load_workbook(excel_saida)

    # --- Formatação para a aba 'Previsao_30_Dias' ---
    if "Previsao_30_Dias" in wb.sheetnames:
        ws_previsao = wb["Previsao_30_Dias"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid") # Verde
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for col_idx, cell in enumerate(ws_previsao[1]): # Itera pelas células do cabeçalho (primeira linha)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            # Auto-ajuste de largura da coluna
            column_letter = get_column_letter(col_idx + 1)
            ws_previsao.column_dimensions[column_letter].width = max(len(str(cell.value)), 15)

        # Formatar coluna 'Data' como Data (coluna C, índice 2)
        # Nota: iter_rows min_col e max_col são baseados em 1, não 0
        for row in ws_previsao.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = 'DD/MM/YYYY' # Formato de data
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

        # AJUSTE FEITO AQUI: Formatar coluna 'Previsão Comissão' como número com 2 casas decimais
        for row in ws_previsao.iter_rows(min_row=2, min_col=4, max_col=4): # Coluna 'Previsão Comissão' é a 4ª (D)
            for cell in row:
                cell.number_format = '#,##0.00' # Formato numérico com duas casas decimais
                cell.alignment = Alignment(horizontal='right')
                cell.border = thin_border

    # --- Formatação para a aba 'Resumo_Explicativo' ---
    if "Resumo_Explicativo" in wb.sheetnames:
        ws_resumo = wb["Resumo_Explicativo"]
        # Aplica a mesma formatação de cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for col_idx, cell in enumerate(ws_resumo[1]):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            column_letter = get_column_letter(col_idx + 1)
            ws_resumo.column_dimensions[column_letter].width = max(len(str(cell.value)), 20) # Ajuste a largura conforme necessário

        # Ajuste de largura específico para a coluna de interpretação
        # Assumindo que "Interpretação da LLM" é a última coluna
        if ws_resumo.max_column >= 5: # Verifica se a coluna existe
            ws_resumo.column_dimensions[get_column_letter(5)].width = 80 # Coluna E, ajustar a largura

    wb.save(excel_saida)
    print("✅ Formatação Excel aplicada com sucesso!")

except Exception as e:
    print(f"⚠️ Erro ao aplicar formatação final ao Excel: {e}")


# === FINAL ===================================================================
# Mensagens de conclusão e resumo do processamento.
print("\n✅ PROCESSAMENTO CONCLUÍDO: Previsão de Comissão")
print("📁 Excel gerado :", excel_saida) # Caminho do arquivo Excel final.
print("🗃️ Gráficos em  :", GRAFICOS_DIR) # Caminho do diretório de gráficos.
print("⏱️ Tempo total  : {:.2f}s".format(time.time() - data_inicio_exec)) # Tempo total de execução.
print("="*60, "\n") # Linha de separação final.